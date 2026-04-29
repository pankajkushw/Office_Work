import openpyxl
import glob
import os
import openpyxl.writer
import openpyxl.writer.excel
import pandas as pd
from pathlib import Path
import shutil


#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
    return rangeSelected

#Takes: start cell, end cell, and sheet you want to copy from.
def copyRangeInternalValue(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).internal_value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
    return rangeSelected

#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

##########################################################################################################
def Swap_A2_for_progress():
    print("Swaping Current Data to Old for progress...")
    wb = openpyxl.load_workbook(report_file, data_only=True) 
    New_sheet = wb['2024-26 A2']
    old_sheet = wb['2024-26 A2_old']
    #copyRange(startCol, startRow, endCol, endRow, sheet)
    copiedData=copyRange(3, 2, 15, 484, New_sheet)
    #pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    pasteRange(3, 2, 15, 484, old_sheet, copiedData)
    wb.save(report_file)
    print("Swaping A2 File completed.")

def Swap_A4_for_progress():
    print("Swaping Current A4 Data to Old for progress...")
    wb = openpyxl.load_workbook(report_file, data_only=True) 
    New_sheet = wb['NewA4Report_2426']
    old_sheet = wb['OldA4Report_2426']
    #copyRange(startCol, startRow, endCol, endRow, sheet)
    copiedData=copyRange(3, 3, 29, 485, New_sheet)
    pasteRange(3, 3, 29, 485, old_sheet, copiedData)
    wb.save(report_file)
    print("Swaping A4 File completed.")


##########################################################################################################
def CopyA2_2425_Data():
    xlfiles = list(RAW_FILE.glob("PhysicalProgressReport*3305*2024-2025*.xlsx"))
    if len(xlfiles) == 0:
        print("No files found in the directory: " + str(RAW_FILE))
    else:
        for x in xlfiles:
            print(x)
    
        #File to be copied
        print("starting A2 Copy")
        #File to be pasted into
        template = openpyxl.load_workbook(report_file, data_only=True)
        # Copying A2 Files 2024-25
        temp_sheet = template["2024-26 A2"] 
        
        for files_list in xlfiles:
            wb = openpyxl.load_workbook(files_list, read_only=True) 
            sheet = wb['Sheet1']
            
            #copyRange(startCol, startRow, endCol, endRow, sheet)
            #pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData)
            files = Path(files_list).name
            print("Copying: "+ files +" in main excel file")
            match files:
                #Bhaiyathan 24-25
                case "PhysicalProgressReport_PMAYG_3305012_2024-2025.xlsx":
                    print("Copying Bhaiyathan 24-25")
                    copiedData=copyRange(1, 3, 13, 80, sheet)
                    pasteRange(31, 2, 43, 79, temp_sheet,copiedData)

                #Odgi 24-25
                case "PhysicalProgressReport_PMAYG_3305013_2024-2025.xlsx":
                    print("Copying Odgi 24-25")
                    copiedData=copyRange(1, 3, 13, 76, sheet)
                    pasteRange(31, 80, 43, 153, temp_sheet,copiedData)
                    
                #Pratappur 24-25
                case "PhysicalProgressReport_PMAYG_3305015_2024-2025.xlsx":
                    print("Copying Pratappur 24-25")
                    copiedData=copyRange(1, 3, 13, 104, sheet)
                    pasteRange(31, 154, 43, 255, temp_sheet,copiedData)
                    
                #Premnagar 24-25
                case "PhysicalProgressReport_PMAYG_3305010_2024-2025.xlsx":
                    print("Copying Premnagar 24-25")
                    copiedData=copyRange(1, 3, 13, 49, sheet)
                    pasteRange(31, 256, 43, 302, temp_sheet,copiedData)
                    
                #Ramanujnagar 24-25
                case "PhysicalProgressReport_PMAYG_3305011_2024-2025.xlsx":
                    print("Copying Ramanujnagar 24-25")
                    copiedData=copyRange(1, 3, 13, 76, sheet)
                    pasteRange(31, 303, 43, 376, temp_sheet,copiedData)
                    
                #Surajpur 24-25
                case "PhysicalProgressReport_PMAYG_3305009_2024-2025.xlsx":
                    print("Copying Surajpur 24-25")
                    copiedData=copyRange(1, 3, 13, 110, sheet)
                    pasteRange(31, 377, 43, 484, temp_sheet,copiedData)
                case _:
                    print("file not matching with any case:" + files)
            wb.close()  
        print("A2 File saved")
        template.save(report_file)
        template.close()
        print("All files copied and pasted successfully")

##########################################################################################################
def CopyA2_2526_Data():
    xlfiles = list(RAW_FILE.glob("PhysicalProgressReport*3305*2025-2026*.xlsx"))
    if len(xlfiles) == 0:
        print("No files found in the directory: " + str(RAW_FILE))
    else:
        for x in xlfiles:
            print(x)
    
        #File to be copied
        print("starting A2 Copy")
        #File to be pasted into
        template = openpyxl.load_workbook(report_file, data_only=True)
        # Copying A2 Files 2024-25
        temp_sheet = template["2025-26 A2"] 
        
        for files_list in xlfiles:
            wb = openpyxl.load_workbook(files_list, read_only=True) 
            sheet = wb['Sheet1']
            
            #copyRange(startCol, startRow, endCol, endRow, sheet)
            #pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData)
            files = Path(files_list).name
            print("Copying: "+ files +" in main excel file")
            match files:
                #Bhaiyathan 24-25
                case "PhysicalProgressReport_PMAYG_3305012_2025-2026.xlsx":
                    print("Copying Bhaiyathan 25-26")
                    copiedData=copyRange(1, 3, 13, 80, sheet)
                    pasteRange(45, 2, 57, 79, temp_sheet,copiedData)

                #Odgi 24-25
                case "PhysicalProgressReport_PMAYG_3305013_2025-2026.xlsx":
                    print("Copying Odgi 25-26")
                    copiedData=copyRange(1, 3, 13, 76, sheet)
                    pasteRange(45, 80, 57, 153, temp_sheet,copiedData)
                    
                #Pratappur 24-25
                case "PhysicalProgressReport_PMAYG_3305015_2025-2026.xlsx":
                    print("Copying Pratappur 25-26")
                    copiedData=copyRange(1, 3, 13, 104, sheet)
                    pasteRange(45, 154, 57, 255, temp_sheet,copiedData)
                    
                #Premnagar 24-25
                case "PhysicalProgressReport_PMAYG_3305010_2025-2026.xlsx":
                    print("Copying Premnagar 25-26")
                    copiedData=copyRange(1, 3, 13, 49, sheet)
                    pasteRange(45, 256, 57, 302, temp_sheet,copiedData)
                    
                #Ramanujnagar 24-25
                case "PhysicalProgressReport_PMAYG_3305011_2025-2026.xlsx":
                    print("Copying Ramanujnagar 25-26")
                    copiedData=copyRange(1, 3, 13, 76, sheet)
                    pasteRange(45, 303, 57, 376, temp_sheet,copiedData)
                    
                #Surajpur 24-25
                case "PhysicalProgressReport_PMAYG_3305009_2025-2026.xlsx":
                    print("Copying Surajpur 25-26")
                    copiedData=copyRange(1, 3, 13, 110, sheet)
                    pasteRange(45, 377, 57, 484, temp_sheet,copiedData)
                case _:
                    print("file not matching with any case:" + files)
            wb.close()  
        print("A2 File saved")
        template.save(report_file)
        template.close()
        print("All files copied and pasted successfully")

##########################################################################################################
def CopyA4_2425_Data():
    xlfiles = list(Path(RAW_FILE).glob("Gap*2024-2025*.xls"))
    print(xlfiles)
    #File to be copied
    print("starting A4 Copy")
    #File to be pasted into
    template = openpyxl.load_workbook(report_file) #Add file name
 
    # Copying A4 Files 2024-25
    print("Opening A4 Sheet")
    temp_sheet = template["NewA4Report_2425"] 
    for files_list in xlfiles:
        
        wb = openpyxl.load_workbook(files_list) 
        sheet = wb['Sheet1']
        
        #copyRange(startCol, startRow, endCol, endRow, sheet)
        #pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData)
        files = Path(files_list).name
        print("Copying: "+ files +" in main excel file")
        match files:
            #Bhaiyathan 24-25
            case "GapinprogressAccountverifctioncompletion_PMAYG_3305012_2024-2025.xlsx":
                print("Copying Bhaiyathan 24-25")
                copiedData=copyRange(1, 3, 27, 81, sheet)
                pasteRange(3, 3, 29, 80, temp_sheet,copiedData)
                                
            #Odgi 24-25
            case "GapinprogressAccountverifctioncompletion_PMAYG_3305013_2024-2025.xlsx":
                print("Copying Odgi 24-25")
                copiedData=copyRange(1, 3, 27, 77, sheet)
                pasteRange(3, 81, 29, 154, temp_sheet,copiedData)
                
            #Pratappur 24-25
            case "GapinprogressAccountverifctioncompletion_PMAYG_3305015_2024-2025.xlsx":
                print("Copying Pratappur 24-25")
                copiedData=copyRange(1, 3, 27, 105, sheet)
                pasteRange(3, 155, 29, 256, temp_sheet,copiedData)
                
            #Premnagar 24-25
            case "GapinprogressAccountverifctioncompletion_PMAYG_3305010_2024-2025.xlsx":
                print("Copying Premnagar 24-25")
                copiedData=copyRange(1, 3, 27, 50, sheet)
                pasteRange(3, 257, 29, 303, temp_sheet,copiedData)
                
            #Ramanujnagar 24-25
            case "GapinprogressAccountverifctioncompletion_PMAYG_3305011_2024-2025.xlsx":
                print("Copying Ramanujnagar 24-25")
                copiedData=copyRange(1, 3, 27, 77, sheet)
                pasteRange(3, 304, 29, 377, temp_sheet,copiedData)
                
            #Surajpur 24-25
            case "GapinprogressAccountverifctioncompletion_PMAYG_3305009_2024-2025.xlsx":
                print("Copying Surajpur 24-25")
                copiedData=copyRange(1, 3, 27, 111, sheet)
                pasteRange(3, 378, 29, 485, temp_sheet,copiedData)
                
            case _:
                print("file not matching with any case:" + files)

        wb.close()  
        print("done")

    print("A4 File saved")
    openpyxl.writer.excel.save_workbook(template, report_file)
    template.close()
        
    print("All files copied and pasted successfully")

##########################################################################################################
def CopyA2_2425_Data_old():

    print("Swaping today's A2 To Old Data")
    #File to be pasted into
    print(report_file)
    template = openpyxl.load_workbook(report_file) #Add file name
    # Copying A2 Files 2024-25
    a2_new = template["2024-24 A2"]
    a2_old = template["2024-24 A2_old"]
    copied_data = copyRange(3, 2, 15, 484, a2_new)
    pasteRange(3, 2, 15, 484, a2_old,copied_data)
    print("Swaping A2 File completed.")
    openpyxl.writer.excel.save_workbook(template, report_file)
    template.save()
    template.close()


##########################################################################################################
def CopyA4_2425_Data_old():
    print("Swaping today's A4 To Old Data")
    #File to be pasted into
    template = openpyxl.load_workbook(report_file) #Add file name
    # Copying A2 Files 2024-25
    a2_new = template["NewA4Report_2425"]
    a2_old = template["OldA4Report_2425"]
    copied_data = copyRange(3, 3, 29, 485, a2_new)
    pasteRange(3, 3, 29, 485, a2_old,copied_data)
    print("Swaping A4 File completed.")
    openpyxl.writer.excel.save_workbook(template, report_file)
    template.close()

##########################################################################################################
def CopyA4_1625_Data_old():
    print("Swaping today's A4 To Old Data")
    #File to be pasted into
    template = openpyxl.load_workbook(report_file, data_only=True) #Add file name
    # Copying A2 Files 2024-25
    a2_new = template["NewA4Report16_23"]
    a2_old = template["OldA4Report16-23"]
    copied_data = copyRange(3, 3, 29, 485, a2_new)
    pasteRange(3, 3, 29, 485, a2_old,copied_data)
    print("Swaping 16-23 A4 in old File completed.")
    openpyxl.writer.excel.save_workbook(template, report_file)
    template.close()    

##########################################################################################################
def CopyA2_1625_Data():
    xlfiles = list(Path(RAW_FILE).glob("Phy*2024-2025*.xls"))
    print(xlfiles)

    #File to be copied
    print("starting A2 16-25 Copy")
    #File to be pasted into
    template = openpyxl.load_workbook(report_file) 
    # Copying A2 Files 1623
    temp_sheet = template["A2_Report_16_23"] 
    for files_list in xlfiles:
        
        wb = openpyxl.load_workbook(files_list) 
        sheet = wb['Sheet1']
        
        #copyRange(startCol, startRow, endCol, endRow, sheet)
        #pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData)
        files = Path(files_list).name
        print("Copying: "+ files +" in main excel file")
        match files:
            #Bhaiyathan 24-25
            case "PhysicalProgressReport_PMAYG_3305012_ALLPMAYG.xlsx":
                print("Copying Bhaiyathan 16-25")
                copiedData=copyRange(1, 3, 13, 80, sheet)
                pasteRange(31, 3, 43, 80, temp_sheet,copiedData)
                                
            #Odgi 24-25
            case "PhysicalProgressReport_PMAYG_3305013_ALLPMAYG.xlsx":
                print("Copying Odgi 16-25")
                copiedData=copyRange(1, 3, 13, 76, sheet)
                pasteRange(31, 81, 43, 154, temp_sheet,copiedData)
                
            #Pratappur 24-25
            case "PhysicalProgressReport_PMAYG_3305015_ALLPMAYG.xlsx":
                print("Copying Pratappur 16-25")
                copiedData=copyRange(1, 3, 13, 104, sheet)
                pasteRange(31, 155,43, 256, temp_sheet,copiedData)
                
            #Premnagar 24-25
            case "PhysicalProgressReport_PMAYG_3305010_ALLPMAYG.xlsx":
                print("Copying Premnagar 16-25")
                copiedData=copyRange(1, 3, 13, 49, sheet)
                pasteRange(31, 257,43, 303, temp_sheet,copiedData)
                
            #Ramanujnagar 24-25
            case "PhysicalProgressReport_PMAYG_3305011_ALLPMAYG.xlsx":
                print("Copying Ramanujnagar 16-25")
                copiedData=copyRange(1, 3, 13, 76, sheet)
                pasteRange(31, 304,43, 377, temp_sheet,copiedData)
                
            #Surajpur 24-25
            case "PhysicalProgressReport_PMAYG_3305009_ALLPMAYG.xlsx":
                print("Copying Surajpur 16-25")
                copiedData=copyRange(1, 3, 13, 110, sheet)
                pasteRange(31, 378,43, 485, temp_sheet,copiedData)
                
            case _:
                print("file not matching with any case:" + files)
  

        wb.close()  
 
    print("A2 File saved")
    openpyxl.writer.excel.save_workbook(template, report_file)
    template.close()
        
    print("All files copied and pasted successfully")

def CopyA4_1625_Data():
    xlfiles = glob.glob( RAW_FILE+"Gap*ALL*.xlsx")
    print(xlfiles)
 
    #File to be copied
    print("starting A4 16-25 Copy")
    #File to be pasted into
    template = openpyxl.load_workbook(report_file) #Add file name
 
    # Copying A4 Files 2024-25
    print("Opening A4 Sheet")
    temp_sheet = template["NewA4Report16_23"] 
    for files_list in xlfiles:
        
        wb = openpyxl.load_workbook(files_list) 
        sheet = wb['Sheet1']
        
        #copyRange(startCol, startRow, endCol, endRow, sheet)
        #pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData)
        files = Path(files_list).name
        print("Copying: "+ files +" in main excel file")
        match files:
            #Bhaiyathan 24-25
            case "GapinprogressAccountverifctioncompletion_PMAYG_3305012_ALLPMAYG.xlsx":
                print("Copying Bhaiyathan 24-25")
                copiedData=copyRange(1, 3, 27, 81, sheet)
                pasteRange(59, 3, 85, 80, temp_sheet,copiedData)
                                
            #Odgi 24-25
            case "GapinprogressAccountverifctioncompletion_PMAYG_3305013_ALLPMAYG.xlsx":
                print("Copying Odgi 24-25")
                copiedData=copyRange(1, 3, 27, 77, sheet)
                pasteRange(59, 81, 85, 154, temp_sheet,copiedData)
                
            #Pratappur 24-25
            case "GapinprogressAccountverifctioncompletion_PMAYG_3305015_ALLPMAYG.xlsx":
                print("Copying Pratappur 24-25")
                copiedData=copyRange(1, 3, 27, 105, sheet)
                pasteRange(59, 155,85, 256, temp_sheet,copiedData)
                
            #Premnagar 24-25
            case "GapinprogressAccountverifctioncompletion_PMAYG_3305010_ALLPMAYG.xlsx":
                print("Copying Premnagar 24-25")
                copiedData=copyRange(1, 3, 27, 50, sheet)
                pasteRange(59, 257,85, 303, temp_sheet,copiedData)
                
            #Ramanujnagar 24-25
            case "GapinprogressAccountverifctioncompletion_PMAYG_3305011_ALLPMAYG.xlsx":
                print("Copying Ramanujnagar 24-25")
                copiedData=copyRange(1, 3, 27, 77, sheet)
                pasteRange(59, 304,85, 377, temp_sheet,copiedData)
                
            #Surajpur 24-25
            case "GapinprogressAccountverifctioncompletion_PMAYG_3305009_ALLPMAYG.xlsx":
                print("Copying Surajpur 24-25")
                copiedData=copyRange(1, 3, 27, 111, sheet)
                pasteRange(59, 378,85, 485, temp_sheet,copiedData)
                
            case _:
                print("file not matching with any case:" + files)
 

        wb.close()  
        print("done")

    print("A4 File saved")
    openpyxl.writer.excel.save_workbook(template, report_file)
    template.close()
    print("All files copied and pasted successfully")

# excution starts here
##########################################################################################################
#dir_path = os.path.dirname(os.path.realpath(__file__))
base_path = Path("D:\\Office\\000Reports\\0000April 2026\\07042026")
raw_file_path = "portalData"
base_file = "PMAYG_TA-AM_WISE_REPORT_06022026_03042026_COLL_M.xlsm"

report_file = base_path.joinpath(base_file)
RAW_FILE = base_path.joinpath(raw_file_path)

print(base_path)
print("raw" + str(RAW_FILE))
print(report_file)

 
file_list_xlsx = list(Path(RAW_FILE).glob("*.xls"))

#Converting xls file into xlsx
for f in file_list_xlsx:
    print(f"Converting: {f}")
    
    # 1. Read the HTML table into a list of DataFrames
    # [0] gets the first table found in the file
    data_list = pd.read_html(f)
    df = data_list[0]
    
    # 2. Create the new filename (switching .xls to .xlsx)
    # Using Path (from pathlib) is cleaner than .replace()
    new_filename = f.with_suffix('.xlsx') 
    
    # 3. Save as a real Excel file
    df.to_excel(new_filename, index=False)
    
    # 4. Remove the old fake .xls file
    shutil.move(f, RAW_FILE.joinpath("backup_folder"))  # Move the original .xls to a backup folder instead of deleting
     
    #os.remove(f)
    print(f"Converted and moved: {f}")


## 24-25 Files
#swap_2425
#Swap_A2_for_progress()
CopyA2_2425_Data()
CopyA2_2526_Data()

#Swap_A4_for_progress()
#CopyA4_2425_Data()
#CopyA4_2526_Data()








    
    

