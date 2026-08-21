import openpyxl
import glob
import os
import openpyxl.writer
import openpyxl.writer.excel
import pandas as pd
from pathlib import Path
import shutil
import time
import warnings

def mySleepFunction(seconds):
    for i in range(seconds):
        print(f"Waiting... {seconds - i} seconds remaining", end="\r")
        time.sleep(1)

#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
    return rangeSelected

#Takes: start cell, end cell, and sheet you want to copy from.
def copyRangeInternalValue(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).internal_value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
    return rangeSelected

#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

##########################################################################################################
def Swap_FTO():
    print("Swaping previous FTO data into backup")
    wb_read = openpyxl.load_workbook(report_file, data_only=True) 
    wb_write = openpyxl.load_workbook(report_file, data_only=False) 
    New_sheet = wb_read['FTO Report_2426']
    old_sheet = wb_write['FTO Report_2426']
    #24-25
    #copyRange(startCol, startRow, endCol, endRow, sheet)
    copiedData=copyRange(7, 7, 8, 13, New_sheet)
    pasteRange(8, 16, 9, 22, old_sheet, copiedData)

    copiedData=copyRange(16, 7, 17, 13, New_sheet)
    pasteRange(16, 16, 17, 22, old_sheet, copiedData)

    copiedData=copyRange(24, 7, 25, 13, New_sheet)
    pasteRange(24, 16, 25, 22, old_sheet, copiedData)

    #25-26
    copiedData=copyRange(7, 29, 8, 35, New_sheet)
    pasteRange(12, 16, 13, 22, old_sheet, copiedData)

    copiedData=copyRange(16, 29, 17, 35, New_sheet)
    pasteRange(20, 16, 21, 22, old_sheet, copiedData)

    copiedData=copyRange(24, 29, 25, 35, New_sheet)
    pasteRange(28, 16, 29, 22, old_sheet, copiedData)
    
    wb_write.save(report_file)
    wb_write.close()
    wb_read.close()
    print("Swaping FTO DATA completed.")
    
def Swap_Daily_Report_Format_1626(report_file):
    print("Step 1: Reading values and formulas securely into memory...")
    
    # Extract calculated text/numbers (Stripping out formulas here)
    wb_data = openpyxl.load_workbook(report_file, data_only=True)
    # Extract raw formulas/structures (This keeps unedited formulas intact)
    wb_raw = openpyxl.load_workbook(report_file, data_only=False)
    
    extracted_values = {}
    
    # Secure local function to copy values into system memory
    def fetch_data(sheet_name, s_col, s_row, e_col, e_row):
        sheet = wb_data[sheet_name]
        matrix = []
        for r in range(s_row, e_row + 1):
            row_cells = []
            for c in range(s_col, e_col + 1):
                row_cells.append(sheet.cell(row=r, column=c).value)
            matrix.append(row_cells)
        return matrix

    # Pulling values out of the data-only reader
    extracted_values['sheet_1626'] = fetch_data('Daily Report Format _1626', 1, 3, 25, 11)
    extracted_values['sheet_2425'] = fetch_data('Daily Report Format24-25', 1, 3, 25, 11)
    
    # Blockwise_new _Sorted extractions
    extracted_values['sorted_plinth'] = fetch_data('Blockwise_new _Sorted', 6, 4, 6, 11)
    extracted_values['sorted_completion'] = fetch_data('Blockwise_new _Sorted', 11, 4, 11, 11)
    extracted_values['sorted_1623'] = fetch_data('Blockwise_new _Sorted', 18, 4, 18, 11)
    
    # Remaining sheets
    extracted_values['sheet_aawas'] = fetch_data('A2_pwl_awas+', 1, 4, 13, 11)
    extracted_values['sheet_combined'] = fetch_data('3_combined 16-23 physical', 7, 3, 7, 10)

    # 6. Blockwise_Rank extraction
    extracted_values['sheet_rank'] = fetch_data('Blockwise_Rank', 10, 3, 10, 11)

    # 7. 4_MR Final extractions
    extracted_values['mr_final_col9'] = fetch_data('4_MR Final', 9, 3, 9, 11)
    extracted_values['mr_final_col12'] = fetch_data('4_MR Final', 12, 3, 12, 11)

    # 8. 4_MR Final extractions
    extracted_values['mmay_row_3'] = fetch_data('5_MMAY', 1, 3, 14, 10)
    extracted_values['mmay_row_12'] = fetch_data('5_MMAY', 1, 12, 16, 19)

    # 9. FTO Data swap
    extracted_values['fto_1_2425'] = fetch_data('FTO Report_2426', 7, 7, 8, 13)
    extracted_values['fto_2_2425'] = fetch_data('FTO Report_2426', 16, 7, 17, 13)
    extracted_values['fto_3_2425'] = fetch_data('FTO Report_2426', 24, 7, 25, 13)
    extracted_values['fto_1_2526'] = fetch_data('FTO Report_2426', 7, 29, 8, 35)
    extracted_values['fto_2_2526'] = fetch_data('FTO Report_2426', 16, 29, 17, 35)
    extracted_values['fto_3_2526'] = fetch_data('FTO Report_2426', 24, 29, 25, 35)
    


    # Close the data instance immediately to release any file hooks
    wb_data.close()

    print("Step 2: Transferring values onto the formula workbook blueprint...")
    
    # Local function to drop pure values onto destinations without corrupting formulas
    def drop_data(sheet_obj, data_matrix, d_col, d_row):
        for r_idx, row_data in enumerate(data_matrix):
            for c_idx, val in enumerate(row_data):
                sheet_obj.cell(row=d_row + r_idx, column=d_col + c_idx, value=val)

    # Drop the text values onto our destinations inside the formula workbook blueprint
    drop_data(wb_raw['Daily Report Format _1626'], extracted_values['sheet_1626'], d_col=30, d_row=3)
    drop_data(wb_raw['Daily Report Format24-25'], extracted_values['sheet_2425'], d_col=30, d_row=3)
    
    sorted_sheet = wb_raw['Blockwise_new _Sorted']
    drop_data(sorted_sheet, extracted_values['sorted_plinth'], d_col=5, d_row=4)
    drop_data(sorted_sheet, extracted_values['sorted_completion'], d_col=10, d_row=4)
    drop_data(sorted_sheet, extracted_values['sorted_1623'], d_col=17, d_row=4)
    
    drop_data(wb_raw['A2_pwl_awas+'], extracted_values['sheet_aawas'], d_col=1, d_row=16)
    drop_data(wb_raw['3_combined 16-23 physical'], extracted_values['sheet_combined'], d_col=6, d_row=3)

    # 6. Blockwise_Rank drop
    drop_data(wb_raw['Blockwise_Rank'], extracted_values['sheet_rank'], d_col=16, d_row=3)

    # 7. 4_MR Final drops
    sheet_mr_final_raw = wb_raw['4_MR Final']
    drop_data(sheet_mr_final_raw, extracted_values['mr_final_col9'], d_col=8, d_row=3)
    drop_data(sheet_mr_final_raw, extracted_values['mr_final_col12'], d_col=11, d_row=3)

    # 8. 5_MMAY drops
    sheet_mmay_raw = wb_raw['5_MMAY']
    drop_data(sheet_mmay_raw, extracted_values['mmay_row_3'], d_col=20, d_row=3)
    drop_data(sheet_mmay_raw, extracted_values['mmay_row_12'], d_col=22, d_row=12)

    # 9. FTO Data swap
    sheet_FTO_raw = wb_raw['FTO Report_2426']
    drop_data(sheet_FTO_raw, extracted_values['fto_1_2425'], d_col=8, d_row=16)
    drop_data(sheet_FTO_raw, extracted_values['fto_2_2425'], d_col=16, d_row=16)
    drop_data(sheet_FTO_raw, extracted_values['fto_3_2425'], d_col=24, d_row=16)
    drop_data(sheet_FTO_raw, extracted_values['fto_1_2526'], d_col=12, d_row=16)
    drop_data(sheet_FTO_raw, extracted_values['fto_2_2526'], d_col=20, d_row=16)
    drop_data(sheet_FTO_raw, extracted_values['fto_3_2526'], d_col=28, d_row=16)




    print("Step 3: Saving modifications...")
    # Saves your changes while keeping all unedited background cells and original formatting completely intact
    wb_raw.save(report_file)
    wb_raw.close()
    print("Success! All data blocks swapped, formulas stripped, and sources preserved.")


##########################################################################################################

COPY_FILE_list = {"A2_2425":"PhysicalProgressReport_PMAYG_3326_2024-2025",
                "A2_2526":"PhysicalProgressReport_PMAYG_3326_2025-2026",
                "A2_1625":"PhysicalProgressReport_PMAYG_3326_ALLPMAYG",
                "A4_2425":"GapinprogressAccountverifctioncompletion_PMAYG_3326_2024-2025",
                "A2_2526":"GapinprogressAccountverifctioncompletion_PMAYG_3326_2025-2026",
                "A2_1625":"GapinprogressAccountverifctioncompletion_PMAYG_3326_ALLPMAYG",
                "F9_2425":"FTOInstallmentWiseReportDetailsNN_Sanctioned_0_0_PMAYG_3326_2024-2025",
                "F9_2526":"FTOInstallmentWiseReportDetailsNN_Sanctioned_0_0_PMAYG_3326_2025-2026",
                "F9_1625":"FTOInstallmentWiseReportDetailsNN_ALLPMAYG_0_0_PMAYG_3326_2025-2026",
                "Awas+_2526":"AwaasPlusPhysicalProgressRprtLogin_District_PMAYG_3326_ALLPMAYG_Cumulative",
                "MMAYG_Ph":"Block_Wise_Physical_Progress_Reports_(Only_Counts)_*",
                "MMAYG_FTO":"FTO_Against_Geotag_Report_*",
                "MR1516":"ongo_comp_pds_wrk_rpt_new",
                "MR1617":"ongo_comp_pds_wrk_rpt_new (1)",
                "MR1718":"ongo_comp_pds_wrk_rpt_new (2)",
                "MR1819":"ongo_comp_pds_wrk_rpt_new (3)",
                "MR1920":"ongo_comp_pds_wrk_rpt_new (4)",
                "MR2021":"ongo_comp_pds_wrk_rpt_new (5)",
                "MR2122":"ongo_comp_pds_wrk_rpt_new (6)",
                "MR2223":"ongo_comp_pds_wrk_rpt_new (7)",
                "MR2324":"ongo_comp_pds_wrk_rpt_new (8)",
                "MR2425":"ongo_comp_pds_wrk_rpt_new (9)",
                "MR2526":"ongo_comp_pds_wrk_rpt_new (10)",
}


def Copy_All_Data():

    xlfiles = backup_file_list_xlsx
    if len(xlfiles) == 0:
        print("No files found in the directory: " + str(backup_file_list_xlsx))
    else:
        for x in xlfiles:
            print(x)
    
        #File to be copied
        print("starting File copy.")
        sheet_write = openpyxl.load_workbook(report_file, data_only=False)
        
        for files_list in xlfiles:
            wb = openpyxl.load_workbook(files_list, read_only=True) 
            sheet = wb['Sheet1']
            
            #copyRange(startCol, startRow, endCol, endRow, sheet)
            #pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData)
            files = Path(files_list).name
            print("Copying: "+ files +" in main excel file")
            match files:
                #24-25
                case "GapinprogressAccountverifctioncompletion_PMAYG_3326_2024-2025.xlsx":
                    print("Copying 24-25")
                    temp_sheet = sheet_write["A4_24-25"] 
                    copiedData=copyRange(1, 2, 27, 9, sheet)
                    pasteRange(1, 4, 27, 11, temp_sheet,copiedData)

                case "GapinprogressAccountverifctioncompletion_PMAYG_3326_2025-2026.xlsx":
                    print("Copying 25-26")
                    temp_sheet = sheet_write["A4_24-25"] 
                    copiedData=copyRange(1, 2, 27, 9, sheet)
                    pasteRange(1, 17, 27, 24, temp_sheet,copiedData)

                # Copying A4 16-26    
                case "GapinprogressAccountverifctioncompletion_PMAYG_3326_ALLPMAYG.xlsx":
                    print("Copying 16-26")
                    temp_sheet = sheet_write["16-26 A4"] 
                    copiedData=copyRange(1, 2, 27, 9, sheet)
                    pasteRange(1, 4, 27, 11, temp_sheet,copiedData)

                # Copying F9 24-26
                case "FTOInstallmentWiseReportDetailsNN_Sanctioned_0_0_PMAYG_3326_2024-2025.xlsx":
                    print("F9 24-25")
                    temp_sheet = sheet_write["F9_24-25"] 
                    copiedData=copyRange(1, 5, 92, 12, sheet)
                    pasteRange(1, 7, 92, 14, temp_sheet,copiedData)

                case "FTOInstallmentWiseReportDetailsNN_Sanctioned_0_0_PMAYG_3326_2025-2026.xlsx":
                    print("F9 25-26")
                    temp_sheet = sheet_write["F9_24-25"]
                    copiedData=copyRange(1, 5, 92, 12, sheet)
                    pasteRange(1, 20, 92, 27, temp_sheet,copiedData)

                case "FTOInstallmentWiseReportDetailsNN_ALLPMAYG_0_0_PMAYG_3326_2025-2026.xlsx":
                    print("F9 25-26")
                    temp_sheet = sheet_write["16-26 F9"]
                    copiedData=copyRange(1, 5, 92, 12, sheet)
                    pasteRange(1, 9, 92, 16, temp_sheet,copiedData)

                case "AwaasPlusPhysicalProgressRprtLogin_District_PMAYG_3326_ALLPMAYG_Cumulative.xlsx":
                    print("A2_pwl_awas+")
                    temp_sheet = sheet_write["A2_pwl_awas+"]
                    copiedData=copyRange(1, 4, 13, 11, sheet)
                    pasteRange(1, 4, 13, 11, temp_sheet,copiedData)                    

                # Copying Physical Report of 24-26
                case "PhysicalProgressReport_PMAYG_3326_2024-2025.xlsx":
                    print("Copying A2 24-25")
                    temp_sheet = sheet_write["A2_pwl_awas+"] 
                    copiedData=copyRange(1, 2, 13, 9, sheet)
                    pasteRange(1, 28, 13, 35, temp_sheet,copiedData)

                case "PhysicalProgressReport_PMAYG_3326_2025-2026.xlsx":
                    print("Copying A2 25-26")
                    temp_sheet = sheet_write["A2_pwl_awas+"] 
                    copiedData=copyRange(1, 2, 13, 9, sheet)
                    pasteRange(1, 39, 13, 46, temp_sheet,copiedData)

                case "PhysicalProgressReport_PMAYG_3326_ALLPMAYG.xlsx":
                    print("Copying A2 24-26")
                    temp_sheet = sheet_write["16-25-A2"] 
                    copiedData=copyRange(1, 2, 13, 9, sheet)
                    pasteRange(1, 3, 13, 10, temp_sheet,copiedData)

                # Gany Report Copy
                case "FTO_Against_Geotag_Report*.xlsx":
                    print("MMAY Physical Report Copy")
                    temp_sheet = sheet_write["A2_pwl_awas+"] 
                    copiedData=copyRange(1, 2, 10, 7, sheet)
                    pasteRange(1, 59, 10, 64, temp_sheet,copiedData)

                case "Block_Wise_Physical_Progress_Reports_(Only_Counts)*.xlsx":
                    print("Copying 25-26")
                    temp_sheet = sheet_write["A2_pwl_awas+"] 
                    copiedData=copyRange(1, 2, 16, 7, sheet)
                    pasteRange(1, 68, 16, 73, temp_sheet,copiedData)

     
                # Copying MR All Year
                case "ongo_comp_pds_wrk_rpt_new.xlsx":
                    print("Copying MR  15-16")
                    temp_sheet = sheet_write["MRAllYear"] 
                    copiedData=copyRange(1, 6, 16, 13, sheet)
                    pasteRange(2, 7, 16, 14, temp_sheet,copiedData)

                case "ongo_comp_pds_wrk_rpt_new (1).xlsx":
                    print("Copying MR  16-17")
                    temp_sheet = sheet_write["MRAllYear"] 
                    copiedData=copyRange(1, 6, 16, 13, sheet)
                    pasteRange(2, 21, 16, 28, temp_sheet,copiedData)

                case "ongo_comp_pds_wrk_rpt_new (2).xlsx":
                    print("Copying MR  17-18")
                    temp_sheet = sheet_write["MRAllYear"] 
                    copiedData=copyRange(1, 6, 16, 13, sheet)
                    pasteRange(2, 35, 16, 42, temp_sheet,copiedData)

                case "ongo_comp_pds_wrk_rpt_new (3).xlsx":
                    print("Copying MR  18-19")
                    temp_sheet = sheet_write["MRAllYear"] 
                    copiedData=copyRange(1, 6, 16, 13, sheet)
                    pasteRange(2, 49, 16, 56, temp_sheet,copiedData)

                case "ongo_comp_pds_wrk_rpt_new (4).xlsx":
                    print("Copying MR  19-20")
                    temp_sheet = sheet_write["MRAllYear"] 
                    copiedData=copyRange(1, 6, 16, 13, sheet)
                    pasteRange(2, 64, 16, 71, temp_sheet,copiedData)

                case "ongo_comp_pds_wrk_rpt_new (5).xlsx":
                    print("Copying MR  20-21")
                    temp_sheet = sheet_write["MRAllYear"] 
                    copiedData=copyRange(1, 6, 16, 13, sheet)
                    pasteRange(2, 78, 16, 85, temp_sheet,copiedData)

                case "ongo_comp_pds_wrk_rpt_new (6).xlsx":
                    print("Copying MR  21-22")
                    temp_sheet = sheet_write["MRAllYear"] 
                    copiedData=copyRange(1, 6, 16, 13, sheet)
                    pasteRange(2, 92, 16, 99, temp_sheet,copiedData)

                case "ongo_comp_pds_wrk_rpt_new (7).xlsx":
                    print("Copying MR  22-23")
                    temp_sheet = sheet_write["MRAllYear"] 
                    copiedData=copyRange(1, 6, 16, 13, sheet)
                    pasteRange(2, 106, 16, 113, temp_sheet,copiedData)

                case "ongo_comp_pds_wrk_rpt_new (8).xlsx":
                    print("Copying MR  23-24")
                    temp_sheet = sheet_write["MRAllYear"] 
                    copiedData=copyRange(1, 6, 16, 13, sheet)
                    pasteRange(2, 120, 16, 127, temp_sheet,copiedData)

                case "ongo_comp_pds_wrk_rpt_new (9).xlsx":
                    print("Copying MR  24-25")
                    temp_sheet = sheet_write["MRAllYear"] 
                    copiedData=copyRange(1, 6, 16, 13, sheet)
                    pasteRange(2, 134, 16, 141, temp_sheet,copiedData)

                case "ongo_comp_pds_wrk_rpt_new (9).xlsx":
                    print("Copying MR  25-26")
                    temp_sheet = sheet_write["MRAllYear"] 
                    copiedData=copyRange(1, 6, 16, 13, sheet)
                    pasteRange(2, 148, 16, 155, temp_sheet,copiedData)
        
            wb.close()  
        print("File saved")
        sheet_write.save(report_file)
        mySleepFunction(2)
        sheet_write.close()

        print("All files copied and pasted successfully")


# excution starts here
##########################################################################################################
base_path = Path("F:\\Office\\000Reports\\0000Aug2026\\21082026")

# 1. Make 'raw_file_path' and 'backup_folder' use Path logic instead of string logic
raw_file_path = Path("portalData")
converted_folder = base_path.joinpath("converted_data") # Joins 'converted_data' to your main directory path

base_file = "PMAYG-ProgressReport_17082026_To_21082026.xlsx"

# 2. Combine the paths properly using Path objects
report_file = base_path.joinpath(base_file)
RAW_FILE = base_path.joinpath(raw_file_path)

# 3. Print statements to check your work
print("Excel files folder is at: " + str(converted_folder))
print("Base path directory:      " + str(base_path))
print("Raw folder path:          " + str(RAW_FILE))
print("Target report file:       " + str(report_file))


warnings.simplefilter("ignore")
file_list_xlsx = list(Path(RAW_FILE).glob("*.xls"))
backup_file_list_xlsx = list(Path(converted_folder).glob("*.xlsx"))
print(f"Number of files found to Convert: {len(file_list_xlsx)}")
print(f"Number of files counverted found: {len(backup_file_list_xlsx)}")

# 4. Converting files into openpyxl readable if not done already.
if len(backup_file_list_xlsx) < len(file_list_xlsx):
    #Converting xls file into xlsx
    for f in file_list_xlsx:
        print(f"Converting: {f}")
        
        # 1. Read the HTML table into a list of DataFrames
        # [0] gets the first table found in the file
        data_list = pd.read_html(f)
        df = data_list[0]
        
        # 2. Create the new filename (switching .xls to .xlsx)
        # Using Path (from pathlib) is cleaner than .replace()
        new_filename = f.with_suffix('.xlsx') 
        
        # 3. Save as a real Excel file
        df.to_excel(new_filename, index=False)
        
        # 4. Remove the old fake .xls file
        if os.path.isdir(converted_folder):
            print("folder exist, moving original files")
        else:
            Path(converted_folder).mkdir(parents=True, exist_ok=True)


        target_backup_file = Path(converted_folder) / f.with_suffix(".xlsx").name

        if target_backup_file.exists():
            print(f"⏭️ Skipping {f.name} (Already converted and moved)")
            continue    
        shutil.move(new_filename, converted_folder)  # Move the original .xls to a backup folder instead of deleting
        #os.remove(f)
        print(f"Converted and moved: {f}, original file left intact.")
else:
    print("Files already converted, skiping convertion...")

# 5. Swaping data of previous date.
#Swap_Daily_Report_Format_1626(report_file)
Copy_All_Data()











    
    

