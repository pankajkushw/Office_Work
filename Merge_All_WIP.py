import openpyxl
import glob
import os
import openpyxl.writer
import openpyxl.writer.excel
import pandas as pd
from pathlib import Path
import shutil
import time
import warnings
import logging 


def mySleepFunction(seconds):
    for i in range(seconds):
        print(f"Waiting... {seconds - i} seconds remaining", end="\r")
        time.sleep(1)

#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
    return rangeSelected

#Takes: start cell, end cell, and sheet you want to copy from.
def copyRangeInternalValue(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).internal_value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
    return rangeSelected

#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

##########################################################################################################

def Swap_MR():
    print("Swaping previous data into backup")
    wb_read = openpyxl.load_workbook(report_file, data_only=True) 
    wb_write = openpyxl.load_workbook(report_file, data_only=False) 
    New_sheet = wb_read['5_MMAY']
    old_sheet = wb_write['5_MMAY']
    #copyRange(startCol, startRow, endCol, endRow, sheet)
    copiedData=copyRange(1, 3, 14, 10, New_sheet)
    pasteRange(20, 3, 33, 10, old_sheet, copiedData)

    copiedData=copyRange(1, 12, 16, 19, New_sheet)
    pasteRange(20, 12, 35, 19, old_sheet, copiedData)
    
    wb_write.save(report_file)
    wb_write.close()
    wb_read.close()
    print("Swaping A4 File completed.")

def Swap_FTO():
    print("Swaping previous FTO data into backup")
    wb_read = openpyxl.load_workbook(report_file, data_only=True) 
    wb_write = openpyxl.load_workbook(report_file, data_only=False) 
    New_sheet = wb_read['FTO Report_2426']
    old_sheet = wb_write['FTO Report_2426']
    #24-25
    #copyRange(startCol, startRow, endCol, endRow, sheet)
    copiedData=copyRange(7, 7, 8, 13, New_sheet)
    pasteRange(8, 16, 9, 22, old_sheet, copiedData)

    copiedData=copyRange(16, 7, 17, 13, New_sheet)
    pasteRange(16, 16, 17, 22, old_sheet, copiedData)

    copiedData=copyRange(24, 7, 25, 13, New_sheet)
    pasteRange(24, 16, 25, 22, old_sheet, copiedData)

    #25-26
    copiedData=copyRange(7, 29, 8, 35, New_sheet)
    pasteRange(12, 16, 13, 22, old_sheet, copiedData)

    copiedData=copyRange(16, 29, 17, 35, New_sheet)
    pasteRange(20, 16, 21, 22, old_sheet, copiedData)

    copiedData=copyRange(24, 29, 25, 35, New_sheet)
    pasteRange(28, 16, 29, 22, old_sheet, copiedData)
    
    wb_write.save(report_file)
    wb_write.close()
    wb_read.close()
    print("Swaping FTO DATA completed.")

def Swap_Daily_Report_Format_1626():
    print("Swaping Daily_Report_Format_1626, 24-25 & Blockwise_new _Sorted & Aawas+ pysical Data")
    wb_read = openpyxl.load_workbook(report_file, data_only=True) 
    wb_write = openpyxl.load_workbook(report_file, data_only=False) 
    New_sheet = wb_read['Daily Report Format _1626']
    old_sheet = wb_write['Daily Report Format _1626']
    #24-25
    #copyRange(startCol, startRow, endCol, endRow, sheet)
    copiedData=copyRange(1, 3, 25, 11, New_sheet)
    pasteRange(30, 3, 54, 11, old_sheet, copiedData)


    New_sheet = wb_read['Daily Report Format24-25']
    old_sheet = wb_write['Daily Report Format24-25']
    #copyRange(startCol, startRow, endCol, endRow, sheet)
    copiedData=copyRange(1, 3, 25, 11, New_sheet)
    pasteRange(30, 3, 54, 11, old_sheet, copiedData)


    New_sheet = wb_read['Blockwise_new _Sorted']
    old_sheet = wb_write['Blockwise_new _Sorted']
    #copyRange(startCol, startRow, endCol, endRow, sheet)
    copiedData=copyRange(6, 4, 6, 11, New_sheet)
    pasteRange(5, 4, 5, 11, old_sheet, copiedData)

    copiedData=copyRange(11, 4, 11, 11, New_sheet)
    pasteRange(10, 4, 10, 11, old_sheet, copiedData)

    copiedData=copyRange(18, 4, 18, 11, New_sheet)
    pasteRange(17, 4, 17, 11, old_sheet, copiedData)

    # Swaping Aawas+ Cummulative Report A2_pwl_awas+
    New_sheet = wb_read['A2_pwl_awas+']
    old_sheet = wb_write['A2_pwl_awas+']
    #copyRange(startCol, startRow, endCol, endRow, sheet)
    copiedData=copyRange(1, 4, 13, 11, New_sheet)
    pasteRange(1, 16, 13, 11, old_sheet, copiedData)

    # Swaping Aawas+ Cummulative Report A2_pwl_awas+
    New_sheet = wb_read['3_combined 16-23 physical']
    old_sheet = wb_write['3_combined 16-23 physical']
    #copyRange(startCol, startRow, endCol, endRow, sheet)
    copiedData=copyRange(7, 3, 7, 10, New_sheet)
    pasteRange(6, 3, 6, 10, old_sheet, copiedData)


    wb_write.save(report_file)
    wb_write.close()
    wb_read.close()
    print("Swaping Daily_Report_Format_1626, 24-25 & Blockwise_new _Sorted completed & Aawas+ pysical Data.")

def Swap_Blockwise_Rank():
    print("Swaping Blockwise_Rank")
    wb_read = openpyxl.load_workbook(report_file, data_only=True) 
    wb_write = openpyxl.load_workbook(report_file, data_only=False) 
    New_sheet = wb_read['Blockwise_Rank']
    old_sheet = wb_write['Blockwise_Rank']
    #24-25
    #copyRange(startCol, startRow, endCol, endRow, sheet)
    copiedData=copyRange(10, 3, 10, 11, New_sheet)
    pasteRange(15, 3, 15, 11, old_sheet, copiedData)
    
    wb_write.save(report_file)
    wb_write.close()
    wb_read.close()
    print("Swaping Blockwise_Rank completed.")    

def Swap_MR():
    print("Swaping MR_Sheet")
    wb_read = openpyxl.load_workbook(report_file, data_only=True) 
    wb_write = openpyxl.load_workbook(report_file, data_only=False) 
    New_sheet = wb_read['4_MR Final']
    old_sheet = wb_write['4_MR Final']
    #24-25
    #copyRange(startCol, startRow, endCol, endRow, sheet)
    copiedData=copyRange(9, 3, 9, 11, New_sheet)
    pasteRange(8, 3, 8, 11, old_sheet, copiedData)

    copiedData=copyRange(12, 3, 12, 11, New_sheet)
    pasteRange(11, 3, 11, 11, old_sheet, copiedData)   

    New_sheet = wb_read['4_MR_Only90Days']
    old_sheet = wb_write['4_MR_Only90Days']
    copiedData=copyRange(6, 3, 6, 11, New_sheet)
    pasteRange(5, 3, 5, 11, old_sheet, copiedData)  

    
    wb_write.save(report_file)
    wb_write.close()
    wb_read.close()
    print("Swaping MR_Sheet completed.")    


##########################################################################################################

COPY_FILE_list = {"A2_2425":"PhysicalProgressReport_PMAYG_3326_2024-2025",
                "A2_2526":"PhysicalProgressReport_PMAYG_3326_2025-2026",
                "A2_1625":"PhysicalProgressReport_PMAYG_3326_ALLPMAYG",
                "A4_2425":"GapinprogressAccountverifctioncompletion_PMAYG_3326_2024-2025",
                "A2_2526":"GapinprogressAccountverifctioncompletion_PMAYG_3326_2025-2026",
                "A2_1625":"GapinprogressAccountverifctioncompletion_PMAYG_3326_ALLPMAYG",
                "F9_2425":"FTOInstallmentWiseReportDetailsNN_Sanctioned_0_0_PMAYG_3326_2024-2025",
                "F9_2526":"FTOInstallmentWiseReportDetailsNN_Sanctioned_0_0_PMAYG_3326_2025-2026",
                "F9_1625":"FTOInstallmentWiseReportDetailsNN_ALLPMAYG_0_0_PMAYG_3326_2025-2026",
                "Awas+_2526":"AwaasPlusPhysicalProgressRprtLogin_District_PMAYG_3326_ALLPMAYG_Cumulative",
                "MMAYG_Ph":"Block_Wise_Physical_Progress_Reports_(Only_Counts)_*",
                "MMAYG_FTO":"FTO_Against_Geotag_Report_*"
}


def Merge_All_files():  

    xlfiles = list(RAW_FILE.glob("PhysicalProgressReport*3305*2024-2025*.xlsx"))
    if len(xlfiles) == 0:
        print("No files found in the directory: " + str(RAW_FILE))
    else:
        for x in xlfiles:
            print(x)
    
        #File to be copied
        print("starting A2 Copy")
        #File to be pasted into
        #template_read = openpyxl.load_workbook(report_file, data_only=True)
        template_write = openpyxl.load_workbook(report_file, data_only=False)
        # Copying A2 Files 2024-25
        temp_sheet = template_write["2024-26 A2"] 
        
        for files_list in xlfiles:
            wb = openpyxl.load_workbook(files_list, read_only=True) 
            sheet = wb['Sheet1']
            
            #copyRange(startCol, startRow, endCol, endRow, sheet)
            #pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData)
            files = Path(files_list).name
            print("Copying: "+ files +" in main excel file")
            match files:
                #Bhaiyathan 24-25
                case "PhysicalProgressReport_PMAYG_3305012_2024-2025.xlsx":
                    print("Copying Bhaiyathan 24-25")
                    copiedData=copyRange(1, 3, 13, 80, sheet)
                    pasteRange(31, 2, 43, 79, temp_sheet,copiedData)

                #Odgi 24-25
                case "PhysicalProgressReport_PMAYG_3305013_2024-2025.xlsx":
                    print("Copying Odgi 24-25")
                    copiedData=copyRange(1, 3, 13, 76, sheet)
                    pasteRange(31, 80, 43, 153, temp_sheet,copiedData)
                    
                #Pratappur 24-25
                case "PhysicalProgressReport_PMAYG_3305015_2024-2025.xlsx":
                    print("Copying Pratappur 24-25")
                    copiedData=copyRange(1, 3, 13, 104, sheet)
                    pasteRange(31, 154, 43, 255, temp_sheet,copiedData)
                    
                #Premnagar 24-25
                case "PhysicalProgressReport_PMAYG_3305010_2024-2025.xlsx":
                    print("Copying Premnagar 24-25")
                    copiedData=copyRange(1, 3, 13, 49, sheet)
                    pasteRange(31, 256, 43, 302, temp_sheet,copiedData)
                    
                #Ramanujnagar 24-25
                case "PhysicalProgressReport_PMAYG_3305011_2024-2025.xlsx":
                    print("Copying Ramanujnagar 24-25")
                    copiedData=copyRange(1, 3, 13, 76, sheet)
                    pasteRange(31, 303, 43, 376, temp_sheet,copiedData)
                    
                #Surajpur 24-25
                case "PhysicalProgressReport_PMAYG_3305009_2024-2025.xlsx":
                    print("Copying Surajpur 24-25")
                    copiedData=copyRange(1, 3, 13, 110, sheet)
                    pasteRange(31, 377, 43, 484, temp_sheet,copiedData)
                case _:
                    print("file not matching with any case:" + files)
            wb.close()  
        print("A2 File saved")
        template_write.save(report_file)
        mySleepFunction(2)
        template_write.close()

        print("All files copied and pasted successfully")
# excution starts here
##########################################################################################################
#dir_path = os.path.dirname(os.path.realpath(__file__)) 
base_path = Path("C:\\Users\\HP\\Downloads\\WIP")
raw_file_path = "portalData"
backup_folder = raw_file_path.join("backup_folder")
base_file = "PMAYG-Meeting_01062026_To_06062026_exceLTL.xlsx"

report_file = base_path.joinpath(base_file)
RAW_FILE = base_path.joinpath(raw_file_path)

print(base_path)
print("raw" + str(RAW_FILE))
print(report_file)

warnings.simplefilter("ignore")
file_list_xlsx = list(Path(RAW_FILE).glob("*.xls"))

#Converting xls file into xlsx
for f in file_list_xlsx:
    print(f"Converting: {f}")
    
    # 1. Read the HTML table into a list of DataFrames
    # [0] gets the first table found in the file
    data_list = pd.read_html(f)
    df = data_list[0]
    
    # 2. Create the new filename (switching .xls to .xlsx)
    # Using Path (from pathlib) is cleaner than .replace()
    new_filename = f.with_suffix('.xlsx') 
    
    # 3. Save as a real Excel file
    df.to_excel(new_filename, index=False)
    
    # 4. Remove the old fake .xls file
    if os.path.isdir(backup_folder):
        print("folder exist, moving original files")
    else:
        Path(backup_folder).mkdir(parents=True, exist_ok=True)
        
    shutil.move(f, RAW_FILE.joinpath("backup_folder"))  # Move the original .xls to a backup folder instead of deleting
    #os.remove(f)
    print(f"Converted and moved: {f}")


Merge_All_files()










    
    

